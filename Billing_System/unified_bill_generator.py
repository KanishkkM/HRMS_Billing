# ================= unified_bill_generator.py =================

import os
import math
import warnings
from datetime import date, timedelta
import pandas as pd
from config import *
from helpers import get_billing_dates
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image as OpenpyxlImage
from num2words import num2words

# Suppress openpyxl WMF image format warning
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.reader.drawings")


# ================= CONFIG PATHS =================
# TEMPLATE_FOLDER and ASSETS_FOLDER are imported from config.py
OUTPUT_FOLDER = OUTPUT_FOLDER  # From config.py

# ================= PO NUMBER MAPPING =================
def load_po_number_mapping():
    """
    Load PO Number data from Excel and create a lookup dictionary
    keyed by (Kind Attention Person, Company Name)
    """
    po_file = "Data/PO_Number.xlsx"
    po_dict = {}
    
    if os.path.exists(po_file):
        try:
            po_df = pd.read_excel(po_file)
            # Clean column names (remove extra spaces)
            po_df.columns = po_df.columns.str.strip()
            
            # Create lookup dictionary
            for _, row in po_df.iterrows():
                kap = str(row.get("Kind Attention Person", "")).strip()
                company = str(row.get("Company", "")).strip()
                key = (kap.lower(), company.lower())
                po_dict[key] = {
                    "PO Number": row.get("PO Number", ""),
                    "Validity": row.get("Validity", "")
                }
            print(f"Loaded {len(po_dict)} PO Number mappings")
        except Exception as e:
            print(f"Warning: Could not load PO Number file: {e}")
    else:
        print(f"Warning: PO Number file not found: {po_file}")
    
    return po_dict


def get_po_details(kap, company, po_dict):
    """
    Get PO Number and Validity for a given KAP and Company
    """
    key = (str(kap).strip().lower(), str(company).strip().lower())
    if key in po_dict:
        return po_dict[key]["PO Number"], po_dict[key]["Validity"]
    return "", ""


# ================= NUMBER TO WORDS =================
def number_to_words_indian(num):
    try:
        words = num2words(num, lang="en_IN")
        return words.title() + " Only"
    except:
        return f"{num} Only"


# ================= BILL PERIOD TEXT =================
def get_billing_period_text(df_group):
    cycle_text = str(df_group.iloc[0]["Billing Cycle"]).lower()

    if "21" in cycle_text and "20" in cycle_text:
        start, end = get_billing_dates("21-20", BILLING_MONTH, BILLING_YEAR)
    elif "25" in cycle_text and "24" in cycle_text:
        start, end = get_billing_dates("25-24", BILLING_MONTH, BILLING_YEAR)
    elif "26" in cycle_text and "25" in cycle_text:
        start, end = get_billing_dates("26-25", BILLING_MONTH, BILLING_YEAR)
    else:
        start, end = get_billing_dates("", BILLING_MONTH, BILLING_YEAR)

    return f"{start.strftime('%d %b %Y')} to {end.strftime('%d %b %Y')}"


# ================= TEMPLATE FILLER =================
def fill_bill_template(ws, totals, gst_values, billing_period_text):
    cgst, sgst, igst, grand_total = gst_values

    today = date.today()
    due_date = today + timedelta(days=3)

    # Dates
    ws["G12"] = today.strftime("%d-%m-%Y")
    ws["G13"] = due_date.strftime("%d-%m-%Y")

    # Billing Period (Merged A17:G20)
    existing_text = ws["A17"].value or ""

    if ":" in existing_text:
        base_text = existing_text.split(":")[0] + ":"
    else:
        base_text = existing_text

    final_text = f"{base_text} {billing_period_text}"

    ws.merge_cells("A17:G20")
    ws["A17"] = final_text

    # Contract Staffing Total (rounded to whole number)
    ws.merge_cells("H17:H20")
    ws["H17"] = round(totals["contract_total"])

    # GST - use math.ceil for CGST/SGST (matching CEILING formula in annexure)
    if igst > 0:
        ws.merge_cells("H21:H22")
        igst_rounded = round(igst)
        ws["H21"] = igst_rounded
        # Grand Total = contract_total + igst
        gt_value = round(totals["contract_total"]) + igst_rounded
    else:
        cgst_ceiled = math.ceil(cgst)
        sgst_ceiled = math.ceil(sgst)
        ws["H21"] = cgst_ceiled
        ws["H22"] = sgst_ceiled
        # Grand Total = contract_total + cgst + sgst
        gt_value = round(totals["contract_total"]) + cgst_ceiled + sgst_ceiled

    # Grand Total
    ws.merge_cells("H23:H25")
    ws["H23"] = gt_value

    # Amount in Words
    ws["B24"] = number_to_words_indian(gt_value)


# ================= FORMAT ANNEXURE SHEET =================
def format_annexure_sheet(ws, num_data_rows, num_cols, annex_columns=None):
    """
    Apply formatting to annexure sheet:
    - Orange header row with increased height for print fit
    - Yellow total row
    - Bold text
    - All borders
    - Whole number display format for numeric columns (hides decimals)
    """
    
    # Non-numeric columns that should NOT get number formatting
    non_numeric_cols = {
        "Kind Attention Person", "Company Name", "Employee Code",
        "Employee Name", "Billing Cycle", "Remark", "Working At",
        "Reporting Person", "Date of Joining"
    }
    
    # Define styles
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Determine which columns are numeric (for whole number display)
    numeric_col_indices = set()
    if annex_columns:
        for col_idx, col_name in enumerate(annex_columns, 1):
            if col_name not in non_numeric_cols:
                numeric_col_indices.add(col_idx)
    
    # Format header row (row 1) - Orange + Bold + Increased height
    ws.row_dimensions[1].height = 30  # Increased header row height for print fit
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = orange_fill
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_alignment
    
    # Format data rows - Borders + whole number format for numeric columns
    for row in range(2, num_data_rows + 2):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col in numeric_col_indices:
                cell.number_format = '0'
    
    # Format total row (last row) - Yellow + Bold + whole number format
    total_row = num_data_rows + 2
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_alignment
        if col in numeric_col_indices:
            cell.number_format = '0'
    
    # Auto-adjust column widths
    for col in range(1, num_cols + 1):
        ws.column_dimensions[chr(64 + col)].width = 18


# ================= ADD IMAGES TO ANNEXURE =================
def add_images_to_annexure(ws, last_row, company_name):
    """
    Add signature and stamp images side by side at bottom of annexure sheet
    - Sign image is selected based on Company Name: sign.png for all except ABNJ (uses sign2.png)
    - Stamp image is selected based on Company Name: Jobuss, Aradhya, or ABNJ
    - Images placed next to each other with increased size
    """
    
    # Determine which sign image to use based on Company Name
    company_lower = str(company_name).lower() if company_name else ""
    
    if "abnj" in company_lower:
        sign_path = os.path.join(ASSETS_FOLDER, "sign2.png")
    else:
        sign_path = os.path.join(ASSETS_FOLDER, "sign.png")
    
    # Determine which stamp image to use based on Company Name
    if "jobuss" in company_lower:
        stamp_filename = "jobuss.png"
    elif "aradhya" in company_lower:
        stamp_filename = "aradhya.png"
    elif "abnj" in company_lower:
        stamp_filename = "abnj.png"
    else:
        # Default to jobuss if no match
        stamp_filename = "jobuss.png"
    
    stamp_path = os.path.join(ASSETS_FOLDER, stamp_filename)
    
    # Position images 2 rows below the total row
    image_row = last_row + 3
    
    # Sign image size
    sign_width = 200
    sign_height = 100
    
    # Stamp image size 125x125
    stamp_width = 125
    stamp_height = 125
    
    try:
        if os.path.exists(sign_path):
            sign_img = OpenpyxlImage(sign_path)
            sign_img.width = sign_width
            sign_img.height = sign_height
            # Place signature at column E (5th column)
            ws.add_image(sign_img, f"E{image_row}")
    except Exception as e:
        print(f"Warning: Could not add signature image: {e}")
    
    try:
        if os.path.exists(stamp_path):
            stamp_img = OpenpyxlImage(stamp_path)
            stamp_img.width = stamp_width
            stamp_img.height = stamp_height
            # Place stamp next to signature at column H (8th column)
            ws.add_image(stamp_img, f"H{image_row}")
        else:
            print(f"Stamp image not found: {stamp_path}")
    except Exception as e:
        print(f"Warning: Could not add stamp image: {e}")


# ================= ADD TOTALS TO ANNEXURE =================
def add_totals_to_annexure(ws, group_df, start_row, annex_columns):
    """
    Add total row with formulas for numeric columns, rounded to 0 digits
    """
    # Columns that should have totals (numeric columns)
    numeric_columns = [
        "Billing", "No of days", "Eligible Days", "No of Saturdays", "No of Sundays", "No of Holidays",
        "Total Present", "Total Working Days", "Absents this Month", "Adjustment of Days",
        "Total Payable Days", "Total Payable Billing", "Charges", "Out of Pocket Exp",
        "Arrears", "Total", "CGST @9%", "SGST @9%", "IGST @18%", "Grand Total"
    ]
    
    num_rows = len(group_df)
    total_row = start_row + num_rows
    
    # Write "TOTAL" in first column
    ws.cell(row=total_row, column=1, value="TOTAL")
    
    # Add formulas for numeric columns, rounded to 0 digits
    # Store column positions for Grand Total calculation
    total_col = None
    cgst_col = None
    sgst_col = None
    igst_col = None
    
    # First pass: identify column positions for all relevant columns
    # Also check if IGST has actual data (non-zero values)
    has_igst_data = False
    for col_idx, col_name in enumerate(annex_columns, 1):
        if col_name in numeric_columns:
            if col_name == "Total":
                total_col = col_idx
            elif col_name == "CGST @9%":
                cgst_col = col_idx
            elif col_name == "SGST @9%":
                sgst_col = col_idx
            elif col_name == "IGST @18%":
                igst_col = col_idx
                # Check if IGST column has any non-zero values in the data
                if col_name in group_df.columns:
                    if group_df[col_name].sum() > 0:
                        has_igst_data = True
    
    # Second pass: create formulas for all numeric columns
    for col_idx, col_name in enumerate(annex_columns, 1):
        if col_name not in numeric_columns:
            continue
            
        col_letter = chr(64 + col_idx)
        
        if col_name == "Grand Total":
            # Calculate Grand Total as sum of Total + applicable GST
            # Use either CGST+SGST OR IGST, not both
            grand_total_formula_parts = []
            if total_col:
                grand_total_formula_parts.append(f"{chr(64 + total_col)}{total_row}")
            
            # Check if IGST is present (column exists and has data)
            has_igst = has_igst_data
            
            if has_igst:
                # Use IGST only (not CGST/SGST)
                if igst_col:
                    grand_total_formula_parts.append(f"{chr(64 + igst_col)}{total_row}")
            else:
                # Use CGST + SGST
                if cgst_col:
                    grand_total_formula_parts.append(f"{chr(64 + cgst_col)}{total_row}")
                if sgst_col:
                    grand_total_formula_parts.append(f"{chr(64 + sgst_col)}{total_row}")
            
            if grand_total_formula_parts:
                formula = f"=ROUND(SUM({','.join(grand_total_formula_parts)}), 0)"
            else:
                formula = f"=ROUND(SUM({col_letter}{start_row}:{col_letter}{total_row - 1}), 0)"
        else:
            if col_name == "CGST @9%" or col_name == "SGST @9%":
                formula = f"=CEILING(SUM({col_letter}{start_row}:{col_letter}{total_row - 1}), 1)"
            else:
                formula = f"=ROUND(SUM({col_letter}{start_row}:{col_letter}{total_row - 1}), 0)"
        
        ws.cell(row=total_row, column=col_idx, value=formula)
    
    return total_row


# ================= MAIN GENERATOR =================
def generate_unified_bills(annex_df):
    """
    Generate unified bills with:
    - Bill sheet (if template exists) + Annexure sheet
    - OR just Annexure sheet (if no template)
    - Master Summary of all annexures
    """
    
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    
    # Load PO Number mapping
    po_dict = load_po_number_mapping()
    
    annex_df["Split_Key"] = (
        annex_df["Kind Attention Person"].astype(str).str.replace(" ", "_")
        + "_"
        + annex_df["Company Name"].astype(str).str.replace(" ", "_")
    )
    
    # Get available templates
    template_files = {
        os.path.splitext(f)[0]: os.path.join(TEMPLATE_FOLDER, f)
        for f in os.listdir(TEMPLATE_FOLDER)
        if f.endswith(".xlsx")
    }
    
    # Store summary data for all annexures
    all_summaries = []
    
    for key, group in annex_df.groupby("Split_Key"):
        
        output_path = os.path.join(OUTPUT_FOLDER, f"{key}.xlsx")
        
        # Remove Split_Key column from output
        group_clean = group.drop(columns=["Split_Key"])
        
        # Remove IGST column if it's not used (all zeros/empty) for this bill
        # This ensures IGST doesn't appear in annexure when CGST/SGST are used
        # Also remove CGST/SGST if IGST is used
        has_igst_data = False
        has_cgst_data = False
        has_sgst_data = False
        
        if "IGST @18%" in group_clean.columns:
            if group_clean["IGST @18%"].sum() > 0:
                has_igst_data = True
                # Remove CGST/SGST if IGST has data
                group_clean = group_clean.drop(columns=["CGST @9%", "SGST @9%"], errors='ignore')
        
        if "CGST @9%" in group_clean.columns:
            if group_clean["CGST @9%"].sum() > 0:
                has_cgst_data = True
        
        if "SGST @9%" in group_clean.columns:
            if group_clean["SGST @9%"].sum() > 0:
                has_sgst_data = True
        
        # If no IGST data, ensure CGST/SGST are used and remove IGST column
        if not has_igst_data and "IGST @18%" in group_clean.columns:
            group_clean = group_clean.drop(columns=["IGST @18%"])
        
        # Check if template exists
        has_template = key in template_files
        
        if has_template:
            # Load template and add annexure
            template_path = template_files[key]
            
            try:
                # Calculate totals for bill from annexure data (2-decimal precision)
                # These values match what the annexure total row will show
                contract_total = group["Total"].sum()
                cgst = group["CGST @9%"].sum()
                sgst = group["SGST @9%"].sum()
                igst = group["IGST @18%"].sum()
                grand_total = group["Grand Total"].sum()
                
                totals = {
                    "contract_total": contract_total,
                    "cgst": cgst,
                    "sgst": sgst,
                    "igst": igst,
                    "grand_total": grand_total
                }
                
                gst_values = (cgst, sgst, igst, grand_total)
                billing_period = get_billing_period_text(group)
                
                # Load template
                wb = load_workbook(template_path)
                bill_sheet = wb.active
                
                # Fill bill template
                fill_bill_template(bill_sheet, totals, gst_values, billing_period)
                
                # Add annexure sheet
                annex_sheet = wb.create_sheet("Annexure")
                
            except Exception as e:
                print(f"Error loading template for {key}: {e}")
                print(f"Creating annexure-only file instead")
                has_template = False
        
        if not has_template:
            # Create new workbook with only annexure
            from openpyxl import Workbook
            wb = Workbook()
            annex_sheet = wb.active
            annex_sheet.title = "Annexure"
        
        # Write annexure data
        # Exclude "Billing Cycle" from annexure output (keep for bill template only)
        # Also exclude Date of Leaving (not needed in output)
        columns_to_exclude = ["Billing Cycle", "Split_Key", "Company Name"]
        annex_columns = [col for col in group_clean.columns if col not in columns_to_exclude]
        
        # Write headers
        for col_idx, header in enumerate(annex_columns, 1):
            annex_sheet.cell(row=1, column=col_idx, value=header)
        
        # Write data rows
        for row_idx, (_, row) in enumerate(group_clean.iterrows(), 2):
            for col_idx, col_name in enumerate(annex_columns, 1):
                annex_sheet.cell(row=row_idx, column=col_idx, value=row[col_name])
        
        # Add totals row
        num_data_rows = len(group_clean)
        total_row = add_totals_to_annexure(annex_sheet, group_clean, 2, annex_columns)
        
        # Format annexure sheet
        num_cols = len(annex_columns)
        format_annexure_sheet(annex_sheet, num_data_rows, num_cols, annex_columns)
        
        # Get company name from the first row of the group
        company_name = group_clean.iloc[0].get("Company Name", "") if len(group_clean) > 0 else ""
        
        # Add images
        add_images_to_annexure(annex_sheet, total_row, company_name)
        
        # Save workbook
        wb.save(output_path)
        
        # Collect summary data - match annexure total row calculations
        # Use CEILING for CGST/SGST (matching annexure formula), ROUND for others
        # Grand Total = Total Amount + applicable GST
        
        total_amount = round(group["Total"].sum(), 2)
        
        # Calculate GST with CEILING for CGST/SGST (matching annexure total row)
        igst_sum = round(group["IGST @18%"].sum(), 2) if "IGST @18%" in group.columns else 0
        cgst_sum = round(group["CGST @9%"].sum(), 2) if "CGST @9%" in group.columns else 0
        sgst_sum = round(group["SGST @9%"].sum(), 2) if "SGST @9%" in group.columns else 0
        
        # Apply CEILING to CGST/SGST (matching annexure total row formula)
        if igst_sum > 0:
            igst_final = round(igst_sum)
            cgst_final = 0
            sgst_final = 0
            grand_total = round(total_amount) + igst_final
        else:
            igst_final = 0
            cgst_final = math.ceil(cgst_sum)
            sgst_final = math.ceil(sgst_sum)
            grand_total = round(total_amount) + cgst_final + sgst_final
        
        summary_data = {
            "Kind Attention Person": group_clean.iloc[0]["Kind Attention Person"] if len(group_clean) > 0 else "",
            "Company Name": company_name,
            "No of Employees": len(group_clean),
            "Total Billing": round(group_clean["Billing"].sum(), 2),
            "Total Payable Billing": round(group_clean["Total Payable Billing"].sum(), 2),
            "Total Charges": round(group_clean["Charges"].sum(), 2),
            "Total Out of Pocket": round(group_clean["Out of Pocket Exp"].sum(), 2),
            "Total Arrears": round(group_clean["Arrears"].sum(), 2),
            "Total Amount": total_amount,
            "CGST": cgst_final,
            "SGST": sgst_final,
            "IGST": igst_final,
            "Grand Total": grand_total
        }
        
        all_summaries.append(summary_data)
        
        status = "with Bill" if has_template else "Annexure only"
        print(f"Generated {key} ({status})")
    
    # Generate Master Summary
    generate_master_summary(all_summaries, po_dict)
    
    print(f"\nAll Unified Bills Generated in '{OUTPUT_FOLDER}' folder")


# ================= MASTER SUMMARY =================
def generate_master_summary(summaries, po_dict):
    """
    Generate a master summary Excel file with all annexure totals
    Includes PO Number and Validity from PO_Number.xlsx
    """
    if not summaries:
        return
    
    from openpyxl import Workbook
    
    summary_path = os.path.join(OUTPUT_FOLDER, "Master_Summary.xlsx")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Summary"
    
    # Define column order: KAP, Company, PO Number, Validity, then numeric columns
    # Order: Total Amount -> GST -> Grand Total
    ordered_headers = [
        "Kind Attention Person",
        "Company Name",
        "PO Number",
        "Validity",
        "No of Employees",
        "Total Billing",
        "Total Payable Billing",
        "Total Charges",
        "Total Out of Pocket",
        "Total Arrears",
        "Total Amount",
    ]
    
    # Always include all GST columns: CGST, SGST, and IGST
    # All summary rows now have all three keys (with 0 for non-applicable)
    ordered_headers.append("CGST")
    ordered_headers.append("SGST")
    ordered_headers.append("IGST")
    
    # Add Grand Total at the end
    ordered_headers.append("Grand Total")
    
    headers = ordered_headers
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[1].height = 30
    
    # Non-numeric columns that should NOT get number formatting
    non_numeric_headers = {"Kind Attention Person", "Company Name", "PO Number", "Validity"}
    
    # Data rows
    for row_idx, summary in enumerate(summaries, 2):
        kap = summary.get("Kind Attention Person", "")
        company = summary.get("Company Name", "")
        
        # Get PO Number and Validity
        po_number, validity = get_po_details(kap, company, po_dict)
        
        for col_idx, header in enumerate(headers, 1):
            if header == "PO Number":
                cell = ws.cell(row=row_idx, column=col_idx, value=po_number)
            elif header == "Validity":
                cell = ws.cell(row=row_idx, column=col_idx, value=validity)
            else:
                # Use .get() to handle missing GST columns gracefully
                cell = ws.cell(row=row_idx, column=col_idx, value=summary.get(header, 0))
            
            # Apply whole number format to numeric columns
            if header not in non_numeric_headers:
                cell.number_format = '0'
    
    # Total row
    total_row = len(summaries) + 2
    ws.cell(row=total_row, column=1, value="GRAND TOTAL")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Sum numeric columns (columns 3 onwards, skipping PO Number and Validity)
    for col_idx in range(3, len(headers) + 1):
        header = headers[col_idx - 1]
        if header in non_numeric_headers:
            continue
        col_letter = chr(64 + col_idx)
        formula = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
        cell = ws.cell(row=total_row, column=col_idx, value=formula)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.number_format = '0'
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    wb.save(summary_path)
    print(f"Master Summary generated: {summary_path}")


# ================= CREATE PLACEHOLDER IMAGES =================
def create_placeholder_images():
    try:
        from PIL import Image, ImageDraw, ImageFont
        
        os.makedirs(ASSETS_FOLDER, exist_ok=True)
        
        # Sign image
        sign_path = os.path.join(ASSETS_FOLDER, "sign.png")
        if not os.path.exists(sign_path):
            img = Image.new('RGB', (300, 150), color='white')
            draw = ImageDraw.Draw(img)
            draw.rectangle([10, 10, 290, 140], outline='black', width=2)
            draw.text((100, 60), "SIGNATURE", fill='black')
            img.save(sign_path)
            print(f"✅ Created placeholder: {sign_path}")
        
        # Stamp images - jobuss, aradhya, abnj
        stamp_types = [
            ("jobuss.png", "JOBUSS", (255, 0, 0)),  # Red
            ("aradhya.png", "ARADHYA", (0, 128, 0)),  # Green
            ("abnj.png", "ABNJ", (0, 0, 255))  # Blue
        ]
        
        for filename, text, color in stamp_types:
            stamp_path = os.path.join(ASSETS_FOLDER, filename)
            if not os.path.exists(stamp_path):
                img = Image.new('RGB', (300, 150), color='white')
                draw = ImageDraw.Draw(img)
                draw.ellipse([10, 10, 290, 140], outline=color, width=3)
                draw.text((100, 60), text, fill=color)
                img.save(stamp_path)
                print(f"Created placeholder: {stamp_path}")
                
    except ImportError:
        print("PIL/Pillow not installed. Please add image files to Assets/ folder manually:")
        print("   - sign.png")
        print("   - jobuss.png")
        print("   - aradhya.png")
        print("   - abnj.png")
    except Exception as e:
        print(f"Warning: Could not create placeholder images: {e}")
