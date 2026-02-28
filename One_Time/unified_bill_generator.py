# ================= unified_bill_generator.py =================
# One_Time Billing - Generates bills with annexures

import os
import warnings
from datetime import date, timedelta
import pandas as pd
from config import *
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image as OpenpyxlImage
from num2words import num2words

# Suppress openpyxl WMF image format warning
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.reader.drawings")


# ================= CONFIG PATHS =================
# TEMPLATE_FOLDER and ASSETS_FOLDER are imported from config.py
OUTPUT_FOLDER = OUTPUT_FOLDER  # From config.py


# ================= NUMBER TO WORDS =================
def number_to_words_indian(num):
    try:
        words = num2words(num, lang="en_IN")
        return words.title() + " Only"
    except:
        return f"{num} Only"


# ================= FILL BILL TEMPLATE =================
def fill_bill_template(ws, group_df):
    """
    Fill the bill template with One_Time billing data:
    - Bill Date: G12
    - Due Date: G13
    - Sum of Charges: H17:H20 (merged)
    - CGST: H21
    - SGST: H22
    - Total (Grand Total): H23:H25 (merged)
    - Rupees(In words): B24
    """
    # Get totals from the group dataframe
    total_charges = round(group_df["Charges"].sum())
    cgst = round(group_df["CGST @9%"].sum())
    sgst = round(group_df["SGST @9%"].sum())
    igst = round(group_df["IGST @18%"].sum())
    grand_total = round(group_df["Grand Total"].sum())
    
    today = date.today()
    due_date = today + timedelta(days=3)
    
    # Dates
    ws["G12"] = today.strftime("%d-%m-%Y")
    ws["G13"] = due_date.strftime("%d-%m-%Y")
    
    # Sum of Charges (merged H17:H20)
    ws.merge_cells("H17:H20")
    ws["H17"] = total_charges
    
    # GST
    if igst > 0:
        ws.merge_cells("H21:H22")
        ws["H21"] = igst
    else:
        ws["H21"] = cgst
        ws["H22"] = sgst
    
    # Grand Total (merged H23:H25)
    ws.merge_cells("H23:H25")
    ws["H23"] = grand_total
    
    # Amount in Words
    ws["B24"] = number_to_words_indian(grand_total)


# ================= FORMAT ANNEXURE SHEET =================
def format_annexure_sheet(ws, num_data_rows, num_cols):
    """
    Apply formatting to annexure sheet:
    - Orange header row with increased height for print fit
    - Yellow total row
    - Bold text
    - All borders
    """
    
    # Define styles
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Format header row (row 1) - Orange + Bold + Increased height
    ws.row_dimensions[1].height = 30
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = orange_fill
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_alignment
    
    # Format data rows - Borders only
    for row in range(2, num_data_rows + 2):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
    
    # Format total row (last row) - Yellow + Bold
    total_row = num_data_rows + 2
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_alignment
    
    # Auto-adjust column widths
    for col in range(1, num_cols + 1):
        ws.column_dimensions[chr(64 + col)].width = 18


# ================= ADD IMAGES TO ANNEXURE =================
def add_images_to_annexure(ws, last_row, company_name):
    """
    Add signature and stamp images side by side at bottom of annexure sheet
    - Sign image is selected based on Company Name: sign.png for all except ABNJ (uses sign2.png)
    - Stamp image is selected based on Company Name: Jobuss, Aradhya, or ABNJ
    """
    # Determine which sign image to use based on Company Name
    company_lower = str(company_name).lower() if company_name else ""
    
    if "abnj" in company_lower:
        sign_path = os.path.join(ASSETS_FOLDER, "sign2.png")
    else:
        sign_path = os.path.join(ASSETS_FOLDER, "sign.png")
    
    # Determine which stamp image to use based on Company Name
    company_lower = str(company_name).lower() if company_name else ""
    
    if "jobuss" in company_lower:
        stamp_filename = "jobuss.png"
    elif "aradhya" in company_lower:
        stamp_filename = "aradhya.png"
    elif "abnj" in company_lower:
        stamp_filename = "abnj.png"
    else:
        stamp_filename = "jobuss.png"
    
    stamp_path = os.path.join(ASSETS_FOLDER, stamp_filename)
    
    # Position images 2 rows below the total row
    image_row = last_row + 3
    
    # Sign image size
    sign_width = 200
    sign_height = 100
    
    # Stamp image size 125x125
    stamp_width = 125
    stamp_height = 125
    
    try:
        if os.path.exists(sign_path):
            sign_img = OpenpyxlImage(sign_path)
            sign_img.width = sign_width
            sign_img.height = sign_height
            ws.add_image(sign_img, f"E{image_row}")
    except Exception as e:
        print(f"Warning: Could not add signature image: {e}")
    
    try:
        if os.path.exists(stamp_path):
            stamp_img = OpenpyxlImage(stamp_path)
            stamp_img.width = stamp_width
            stamp_img.height = stamp_height
            ws.add_image(stamp_img, f"H{image_row}")
        else:
            print(f"Stamp image not found: {stamp_path}")
    except Exception as e:
        print(f"Warning: Could not add stamp image: {e}")


# ================= ADD TOTALS TO ANNEXURE =================
def add_totals_to_annexure(ws, group_df, start_row, annex_columns):
    """
    Add total row with formulas for numeric columns
    """
    # Columns that should have totals (numeric columns)
    numeric_columns = [
        "Charges", "Total", "CGST @9%", "SGST @9%", "IGST @18%", "Grand Total"
    ]
    
    num_rows = len(group_df)
    total_row = start_row + num_rows
    
    # Write "TOTAL" in first column
    ws.cell(row=total_row, column=1, value="TOTAL")
    
    # Store column positions for Grand Total calculation
    total_col = None
    cgst_col = None
    sgst_col = None
    igst_col = None
    
    # Add formulas for numeric columns
    for col_idx, col_name in enumerate(annex_columns, 1):
        if col_name in numeric_columns:
            col_letter = chr(64 + col_idx)
            
            if col_name == "Grand Total":
                # Calculate Grand Total as sum of Total + CGST + SGST + IGST
                grand_total_formula_parts = []
                if total_col:
                    grand_total_formula_parts.append(f"{chr(64 + total_col)}{total_row}")
                if cgst_col:
                    grand_total_formula_parts.append(f"{chr(64 + cgst_col)}{total_row}")
                if sgst_col:
                    grand_total_formula_parts.append(f"{chr(64 + sgst_col)}{total_row}")
                if igst_col:
                    grand_total_formula_parts.append(f"{chr(64 + igst_col)}{total_row}")
                
                if grand_total_formula_parts:
                    formula = f"=ROUND(SUM({','.join(grand_total_formula_parts)}), 0)"
                else:
                    formula = f"=ROUND(SUM({col_letter}{start_row}:{col_letter}{total_row - 1}), 0)"
            else:
                if col_name == "CGST @9%" or col_name == "SGST @9%":
                    formula = f"=CEILING(SUM({col_letter}{start_row}:{col_letter}{total_row - 1}), 1)"
                else:
                    formula = f"=ROUND(SUM({col_letter}{start_row}:{col_letter}{total_row - 1}), 0)"
                
                # Store column positions for Grand Total calculation
                if col_name == "Total":
                    total_col = col_idx
                elif col_name == "CGST @9%":
                    cgst_col = col_idx
                elif col_name == "SGST @9%":
                    sgst_col = col_idx
                elif col_name == "IGST @18%":
                    igst_col = col_idx
            
            ws.cell(row=total_row, column=col_idx, value=formula)
    
    return total_row


# ================= MAIN GENERATOR =================
def generate_unified_bills(annex_df):
    """
    Generate unified bills with:
    - Bill sheet (if template exists) + Annexure sheet
    - OR just Annexure sheet (if no template)
    - Master Summary of all annexures
    """
    
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    
    annex_df["Split_Key"] = (
        annex_df["Kind Attention Person"].astype(str).str.replace(" ", "_")
        + "_"
        + annex_df["Company Name"].astype(str).str.replace(" ", "_")
    )
    
    # Get available templates
    template_files = {
        os.path.splitext(f)[0]: os.path.join(TEMPLATE_FOLDER, f)
        for f in os.listdir(TEMPLATE_FOLDER)
        if f.endswith(".xlsx")
    }
    
    # Store summary data for all annexures
    all_summaries = []
    
    for key, group in annex_df.groupby("Split_Key"):
        
        output_path = os.path.join(OUTPUT_FOLDER, f"{key}_OneTime.xlsx")
        
        # Remove Split_Key column from output
        group_clean = group.drop(columns=["Split_Key"])
        
        # Check if template exists
        has_template = key in template_files
        
        if has_template:
            # Load template and add annexure
            template_path = template_files[key]
            
            try:
                wb = load_workbook(template_path)
                # Get the first sheet (bill sheet) and fill it with data
                bill_sheet = wb.active
                fill_bill_template(bill_sheet, group_clean)
                
                # Create annexure sheet
                annex_sheet = wb.create_sheet("Annexure")
            except Exception as e:
                print(f"Error loading template for {key}: {e}")
                print(f"Creating annexure-only file instead")
                has_template = False
        
        if not has_template:
            # Create new workbook with only annexure
            wb = Workbook()
            annex_sheet = wb.active
            annex_sheet.title = "Annexure"
        
        # Write annexure data
        # Exclude Company Name from detailed output (keep for reference)
        columns_to_exclude = ["Split_Key", "Company Name"]
        annex_columns = [col for col in group_clean.columns if col not in columns_to_exclude]
        
        # Write headers
        for col_idx, header in enumerate(annex_columns, 1):
            annex_sheet.cell(row=1, column=col_idx, value=header)
        
        # Write data rows
        for row_idx, (_, row) in enumerate(group_clean.iterrows(), 2):
            for col_idx, col_name in enumerate(annex_columns, 1):
                annex_sheet.cell(row=row_idx, column=col_idx, value=row[col_name])
        
        # Add totals row
        num_data_rows = len(group_clean)
        total_row = add_totals_to_annexure(annex_sheet, group_clean, 2, annex_columns)
        
        # Format annexure sheet
        num_cols = len(annex_columns)
        format_annexure_sheet(annex_sheet, num_data_rows, num_cols)
        
        # Get company name from the first row of the group
        company_name = group_clean.iloc[0].get("Company Name", "") if len(group_clean) > 0 else ""
        
        # Add images
        add_images_to_annexure(annex_sheet, total_row, company_name)
        
        # Save workbook
        wb.save(output_path)
        
        # Collect summary data - dynamically handle GST columns
        summary_data = {
            "Kind Attention Person": group_clean.iloc[0]["Kind Attention Person"] if len(group_clean) > 0 else "",
            "Working At": group_clean.iloc[0]["Working At"] if len(group_clean) > 0 else "",
            "Company Name": company_name,
            "No of Employees": len(group_clean),
            "Total Charges": round(group_clean["Charges"].sum()),
            "Total Amount": round(group_clean["Total"].sum()),
            "Grand Total": round(group_clean["Grand Total"].sum())
        }
        
        # Add only the applicable GST columns
        if "IGST @18%" in group_clean.columns:
            summary_data["IGST"] = round(group_clean["IGST @18%"].sum())
        else:
            if "CGST @9%" in group_clean.columns:
                summary_data["CGST"] = round(group_clean["CGST @9%"].sum())
            if "SGST @9%" in group_clean.columns:
                summary_data["SGST"] = round(group_clean["SGST @9%"].sum())
        
        all_summaries.append(summary_data)
        
        status = "with Bill" if has_template else "Annexure only"
        print(f"Generated {key} ({status})")
    
    # Generate Master Summary
    generate_master_summary(all_summaries)
    
    print(f"\nAll One_Time Bills Generated in '{OUTPUT_FOLDER}' folder")


# ================= MASTER SUMMARY =================
def generate_master_summary(summaries):
    """
    Generate a master summary Excel file with all annexure totals
    """
    if not summaries:
        return
    
    summary_path = os.path.join(OUTPUT_FOLDER, SUMMARY_FILE)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Summary"
    
    # Headers
    headers = list(summaries[0].keys())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[1].height = 30
    
    # Data rows
    for row_idx, summary in enumerate(summaries, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=summary[header])
    
    # Total row
    total_row = len(summaries) + 2
    ws.cell(row=total_row, column=1, value="GRAND TOTAL")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Sum numeric columns (columns 3 onwards)
    for col_idx in range(3, len(headers) + 1):
        col_letter = chr(64 + col_idx)
        formula = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
        cell = ws.cell(row=total_row, column=col_idx, value=formula)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    wb.save(summary_path)
    print(f"Master Summary generated: {summary_path}")


# ================= CREATE PLACEHOLDER IMAGES =================
def create_placeholder_images():
    try:
        from PIL import Image, ImageDraw, ImageFont
        
        os.makedirs(ASSETS_FOLDER, exist_ok=True)
        
        # Sign image
        sign_path = os.path.join(ASSETS_FOLDER, "sign.png")
        if not os.path.exists(sign_path):
            img = Image.new('RGB', (300, 150), color='white')
            draw = ImageDraw.Draw(img)
            draw.rectangle([10, 10, 290, 140], outline='black', width=2)
            draw.text((100, 60), "SIGNATURE", fill='black')
            img.save(sign_path)
            print(f"Created placeholder: {sign_path}")
        
        # Stamp images - jobuss, aradhya, abnj
        stamp_types = [
            ("jobuss.png", "JOBUSS", (255, 0, 0)),  # Red
            ("aradhya.png", "ARADHYA", (0, 128, 0)),  # Green
            ("abnj.png", "ABNJ", (0, 0, 255))  # Blue
        ]
        
        for filename, text, color in stamp_types:
            stamp_path = os.path.join(ASSETS_FOLDER, filename)
            if not os.path.exists(stamp_path):
                img = Image.new('RGB', (300, 150), color='white')
                draw = ImageDraw.Draw(img)
                draw.ellipse([10, 10, 290, 140], outline=color, width=3)
                draw.text((100, 60), text, fill=color)
                img.save(stamp_path)
                print(f"Created placeholder: {stamp_path}")
                
    except ImportError:
        print("PIL/Pillow not installed. Please add image files to Assets/ folder manually:")
